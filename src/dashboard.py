"""JobHunter 数据看板模块。

汇总 data/ 下所有 JSON 数据，生成 Excel 看板导出。
"""
from __future__ import annotations

import json
import sys
from typing import Any, Dict, List

from . import config


def load_jobs() -> List[Dict[str, Any]]:
    jobs_path = config.DATA_DIR / "jobs.json"
    if not jobs_path.exists():
        return []
    with open(jobs_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_messages() -> List[Dict[str, Any]]:
    messages_path = config.DATA_DIR / "messages.json"
    if not messages_path.exists():
        return []
    with open(messages_path, "r", encoding="utf-8") as f:
        return json.load(f)


def run() -> None:
    """主入口：生成 Excel 看板。"""
    config.ensure_dirs()
    jobs = load_jobs()
    messages = load_messages()

    if not jobs and not messages:
        print("暂无求职数据，请先执行采集。")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # Sheet 1: 岗位总览
    ws1 = wb.active
    ws1.title = "岗位总览"
    headers = ["职位", "公司", "薪资", "城市", "评分", "评分理由", "状态", "投递时间"]
    ws1.append(headers)
    for c in ws1[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    for j in jobs:
        ws1.append([
            j.get("title", ""),
            j.get("company", ""),
            j.get("salary", ""),
            j.get("city", ""),
            j.get("score", ""),
            j.get("score_reason", ""),
            j.get("status", ""),
            j.get("sent_at", ""),
        ])

    # Sheet 2: 消息记录
    ws2 = wb.create_sheet("消息记录")
    ws2.append(["公司", "最后消息", "时间", "抓取时间"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for m in messages:
        ws2.append([
            m.get("company", ""),
            m.get("last_msg", ""),
            m.get("time", ""),
            m.get("fetched_at", ""),
        ])

    # Sheet 3: 统计
    ws3 = wb.create_sheet("统计")
    status_counts: Dict[str, int] = {}
    for j in jobs:
        s = j.get("status", "未知")
        status_counts[s] = status_counts.get(s, 0) + 1
    ws3.append(["指标", "数值"])
    ws3.append(["岗位总数", len(jobs)])
    ws3.append(["消息总数", len(messages)])
    for s, c in status_counts.items():
        ws3.append([f"状态-{s}", c])
    avg_score = sum(j.get("score", 0) for j in jobs) / len(jobs) if jobs else 0
    ws3.append(["平均评分", round(avg_score, 1)])

    out_path = config.EXPORT_DIR / "求职看板.xlsx"
    wb.save(out_path)
    print(f"求职看板已生成：{out_path}")
    print(f"岗位 {len(jobs)} 个，消息 {len(messages)} 条。")


if __name__ == "__main__":
    try:
        run()
    except Exception as e:
        print(f"看板生成出错：{e}")
        sys.exit(1)
